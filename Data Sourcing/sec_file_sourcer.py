import pandas as pd
import requests
import json
from datetime import datetime, timedelta
import time
from typing import Dict, List, Optional, Tuple
import warnings
import os
import sys
from difflib import SequenceMatcher
import re
import numpy as np
warnings.filterwarnings('ignore')

from joblib import Parallel, delayed
try:
    from rapidfuzz import fuzz, process as rf_process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

class SECFileSourcer:
    """
    A comprehensive class for sourcing SEC filings and creating financial models.
    
    This class provides methods to:
    1. Find and download SEC 10-K and 10-Q filings
    2. Create traditional financial three-statement models
    3. Generate sensitivity analysis with operating leverage impacts
    4. Create KPI summary sheets
    """
    
    # Class-level cache for ticker-to-CIK mapping
    _ticker_cik_cache = None
    _cache_last_updated = None
    _cache_expiry_seconds = 24 * 3600  # 24 hours
    _fuzzy_match_cache = {}
    
    def __init__(self):
        """Initialize the SEC File Sourcer with base URLs and headers."""
        self.base_url = "https://data.sec.gov"
        self.sec_ticker_url = "https://www.sec.gov/files/company_tickers.json"
        self.headers = {
            'User-Agent': 'ModelMaker/1.0 (kit.kumar@gmail.com)',
            'Accept': 'application/json',
            'Accept-Encoding': 'gzip, deflate'
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        
        # Add rate limiting - SEC requires delays between requests
        self.last_request_time = 0
        self.min_request_interval = 0.1  # 100ms minimum between requests
        
    def _rate_limit(self):
        """Ensure minimum time between API requests to comply with SEC rate limits."""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        if time_since_last < self.min_request_interval:
            time.sleep(self.min_request_interval - time_since_last)
        self.last_request_time = time.time()

    @classmethod
    def _is_cache_valid(cls):
        if cls._ticker_cik_cache is None or cls._cache_last_updated is None:
            return False
        return (time.time() - cls._cache_last_updated) < cls._cache_expiry_seconds

    @classmethod
    def _load_ticker_cik_cache(cls, session, url):
        """Load the ticker-to-CIK mapping from the SEC and cache it."""
        try:
            response = session.get(url)
            if response.status_code != 200:
                print(f"Error fetching company tickers: {response.status_code}")
                return None
            data = response.json()
            # The SEC file is now a dictionary with numeric string keys
            cache = {}
            for key, entry in data.items():
                if isinstance(entry, dict) and 'ticker' in entry and 'cik_str' in entry:
                    cache[entry['ticker'].upper()] = str(entry['cik_str']).zfill(10)
            cls._ticker_cik_cache = cache
            cls._cache_last_updated = time.time()
            return cache
        except Exception as e:
            print(f"Error loading ticker-to-CIK cache: {e}")
            return None

    def get_cik_from_ticker(self, ticker: str, force_refresh: bool = False) -> Optional[str]:
        """
        Convert a stock ticker symbol to its corresponding CIK (Central Index Key) number.
        Uses a cached mapping for efficiency.
        Args:
            ticker (str): Stock ticker symbol (e.g., 'AAPL', 'MSFT')
            force_refresh (bool): If True, refresh the cache from the SEC
        Returns:
            Optional[str]: CIK number as a string (10-digit zero-padded), or None if not found
        """
        ticker_upper = ticker.upper()
        # Check cache
        if force_refresh or not self._is_cache_valid():
            self._rate_limit()
            cache = self._load_ticker_cik_cache(self.session, self.sec_ticker_url)
            if cache is None:
                print("Could not fetch ticker-to-CIK mapping from SEC.")
                return None
        else:
            cache = self._ticker_cik_cache
        cik = cache.get(ticker_upper)
        if cik:
            return cik
        print(f"Ticker '{ticker}' not found in SEC database.")
        return None
    
    def find_sec_filings(self, ticker: str, filing_types: List[str] = ['10-K', '10-Q']) -> pd.DataFrame:
        """
        Find SEC 10-K and 10-Q filings for a given stock ticker, sorted by date.
        
        Args:
            ticker (str): Stock ticker symbol
            filing_types (List[str]): Types of filings to search for (default: ['10-K', '10-Q'])
            
        Returns:
            pd.DataFrame: DataFrame containing filing information sorted by date
        """
        try:
            # Convert ticker to CIK
            cik = self.get_cik_from_ticker(ticker)
            if not cik:
                print(f"Could not find CIK for ticker: {ticker}")
                return pd.DataFrame()
            
            # Get company submissions
            submissions_url = f"{self.base_url}/submissions/CIK{cik}.json"
            response = self.session.get(submissions_url)
            
            if response.status_code != 200:
                print(f"Error fetching submissions for {ticker} (CIK: {cik}): {response.status_code}")
                return pd.DataFrame()
            
            submissions_data = response.json()
            filings = submissions_data.get('filings', {}).get('recent', {})
            
            # Create DataFrame from filings
            filing_data = []
            for i in range(len(filings.get('form', []))):
                form = filings['form'][i]
                if form in filing_types:
                    filing_data.append({
                        'form': form,
                        'filingDate': filings['filingDate'][i],
                        'accessionNumber': filings['accessionNumber'][i],
                        'primaryDocument': filings['primaryDocument'][i],
                        'description': filings.get('description', [''])[i] if i < len(filings.get('description', [])) else ''
                    })
            
            df = pd.DataFrame(filing_data)
            if not df.empty:
                df['filingDate'] = pd.to_datetime(df['filingDate'])
                df = df.sort_values('filingDate', ascending=False)
                
            return df
            
        except Exception as e:
            print(f"Error in find_sec_filings: {str(e)}")
            return pd.DataFrame()
    
    def create_financial_model(self, ticker: str, quarters: int = 8, progress_callback=None, enhanced_fuzzy_matching: bool = True) -> Dict[str, pd.DataFrame]:
        """
        Create a traditional financial three-statement model with annual and quarterly views.
        
        Args:
            ticker (str): Stock ticker symbol
            quarters (int): Number of quarters of data to retrieve (default: 8, which is 2 years)
                          This will determine how many 10-K and 10-Q filings to process.
                          For example: 4 quarters = 1 year, 8 quarters = 2 years, 12 quarters = 3 years
            progress_callback (callable): Optional callback function for progress updates
            enhanced_fuzzy_matching (bool): Whether to include non-GAAP to GAAP mapping (default: True)
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing financial model dataframes
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            # Store the quarters parameter for use in data filtering
            self.current_quarters = quarters
            
            # Initialize the financial model structure
            financial_model = {
                'annual_income_statement': pd.DataFrame(),
                'annual_balance_sheet': pd.DataFrame(),
                'annual_cash_flow': pd.DataFrame(),
                'quarterly_income_statement': pd.DataFrame(),
                'quarterly_balance_sheet': pd.DataFrame(),
                'quarterly_cash_flow': pd.DataFrame()
            }
            
            # Convert ticker to CIK
            progress("    • Converting ticker to CIK number...")
            cik = self.get_cik_from_ticker(ticker)
            if not cik:
                progress(f"    ✗ Could not find CIK for ticker: {ticker}")
                return financial_model
            
            progress(f"    ✓ Found CIK: {cik}")
            
            # Remove leading zeros from CIK for new endpoint
            cik_no_zeros = str(int(cik))
            # Get company facts from new endpoint
            progress("    • Fetching company facts from SEC API...")
            company_facts_url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik_no_zeros}.json"
            response = self.session.get(company_facts_url)
            
            if response.status_code == 200:
                progress("    ✓ Successfully retrieved company facts from SEC API")
                facts_data = response.json()
                facts = facts_data.get('facts', {})
                # Define key financial metrics for each statement
                income_statement_metrics = {
                    'Revenues': 'us-gaap:Revenues',
                    'CostOfRevenue': 'us-gaap:CostOfRevenue',
                    'GrossProfit': 'us-gaap:GrossProfit',
                    'OperatingExpenses': 'us-gaap:OperatingExpenses',
                    'OperatingIncomeLoss': 'us-gaap:OperatingIncomeLoss',
                    'NetIncomeLoss': 'us-gaap:NetIncomeLoss',
                    'EarningsPerShareBasic': 'us-gaap:EarningsPerShareBasic',
                    'EarningsPerShareDiluted': 'us-gaap:EarningsPerShareDiluted'
                }
                balance_sheet_metrics = {
                    'CashAndCashEquivalents': 'us-gaap:CashAndCashEquivalentsAtCarryingValue',
                    'TotalAssets': 'us-gaap:Assets',
                    'TotalCurrentAssets': 'us-gaap:AssetsCurrent',
                    'TotalLiabilities': 'us-gaap:Liabilities',
                    'TotalCurrentLiabilities': 'us-gaap:LiabilitiesCurrent',
                    'TotalStockholdersEquity': 'us-gaap:StockholdersEquity',
                    'TotalDebt': 'us-gaap:LongTermDebtNoncurrent'
                }
                cash_flow_metrics = {
                    'NetCashProvidedByUsedInOperatingActivities': 'us-gaap:NetCashProvidedByUsedInOperatingActivities',
                    'NetCashProvidedByUsedInInvestingActivities': 'us-gaap:NetCashProvidedByUsedInInvestingActivities',
                    'NetCashProvidedByUsedInFinancingActivities': 'us-gaap:NetCashProvidedByUsedInFinancingActivities',
                    'CapitalExpenditures': 'us-gaap:PaymentsToAcquirePropertyPlantAndEquipment',
                    'DividendsPaid': 'us-gaap:PaymentsOfDividends'
                }
                
                progress("    • Extracting financial data from API response...")
                for period in ['annual', 'quarterly']:
                    for statement, metrics in [('income_statement', income_statement_metrics),
                                             ('balance_sheet', balance_sheet_metrics),
                                             ('cash_flow', cash_flow_metrics)]:
                        df = self._extract_financial_data(facts, metrics, period)
                        financial_model[f'{period}_{statement}'] = df
                
                progress("    ✓ Financial model created from SEC API data")
                return financial_model
            else:
                progress(f"    ⚠ SEC API not available (status: {response.status_code}), falling back to XBRL parsing...")
                # Fallback: Try to find and parse XBRL XML instance document using Arelle
                try:
                    progress("    • Finding SEC filings for XBRL extraction...")
                    filings_df = self.find_sec_filings(ticker)
                    if filings_df.empty:
                        progress(f"    ✗ No filings found for {ticker}")
                        return financial_model
                    
                    # Calculate how many filings we need based on quarters parameter
                    years_needed = max(1, (quarters + 3) // 4)  # Round up to get full years needed
                    k_filings_needed = years_needed  # One 10-K per year
                    q_filings_needed = min(quarters, 20)  # Limit 10-Q filings to prevent excessive processing
                    
                    progress(f"    • Requested {quarters} quarters of data ({years_needed} years)")
                    progress(f"    • Will process up to {k_filings_needed} 10-K filings and {q_filings_needed} 10-Q filings")
                    
                    # Separate 10-K and 10-Q filings
                    k_filings = filings_df[filings_df['form'] == '10-K'].head(k_filings_needed)
                    q_filings = filings_df[filings_df['form'] == '10-Q'].head(q_filings_needed)
                    
                    progress(f"    ✓ Found {len(k_filings)} 10-K filings and {len(q_filings)} 10-Q filings")
                    progress("    • Starting comprehensive XBRL data extraction...")
                    
                    # Initialize comprehensive financial model
                    comprehensive_facts = {}
                    annual_data = {}
                    quarterly_data = {}
                    
                    # Process 10-K filings first (primary source for annual data)
                    progress(f"    • Processing {len(k_filings)} 10-K filings (annual data)...")
                    for i, (idx, row) in enumerate(k_filings.iterrows(), 1):
                        progress(f"      [{i}/{len(k_filings)}] Processing 10-K: {row['filingDate']}")
                        k_facts = self._extract_xbrl_from_filing(row, cik, progress_callback)
                        if k_facts:
                            # Store annual data from 10-K
                            for concept, data in k_facts.items():
                                if concept not in annual_data:
                                    annual_data[concept] = []
                                annual_data[concept].extend(data)
                            comprehensive_facts.update(k_facts)
                            progress(f"        ✓ Extracted {len(k_facts)} concepts")
                        else:
                            progress(f"        ✗ No data extracted")
                    
                    # Process 10-Q filings (supplementary quarterly data)
                    progress(f"    • Processing {len(q_filings)} 10-Q filings (quarterly data)...")
                    for i, (idx, row) in enumerate(q_filings.iterrows(), 1):
                        progress(f"      [{i}/{len(q_filings)}] Processing 10-Q: {row['filingDate']}")
                        q_facts = self._extract_xbrl_from_filing(row, cik, progress_callback)
                        if q_facts:
                            # Store quarterly data from 10-Q
                            for concept, data in q_facts.items():
                                if concept not in quarterly_data:
                                    quarterly_data[concept] = []
                                quarterly_data[concept].extend(data)
                            comprehensive_facts.update(q_facts)
                            progress(f"        ✓ Extracted {len(q_facts)} concepts")
                        else:
                            progress(f"        ✗ No data extracted")
                    
                    if not comprehensive_facts:
                        progress("    ✗ No XBRL data found in any filings")
                        return financial_model
                    
                    progress(f"    ✓ Comprehensive data extracted: {len(comprehensive_facts)} unique concepts")
                    progress(f"    • Annual data: {sum(len(data) for data in annual_data.values())} data points")
                    progress(f"    • Quarterly data: {sum(len(data) for data in quarterly_data.values())} data points")
                    
                    # Run discrepancy checks between annual and quarterly data
                    progress("    • Running data consistency checks...")
                    self._run_discrepancy_checks(annual_data, quarterly_data)
                    
                    # Create financial model using comprehensive data
                    progress("    • Creating comprehensive financial model...")
                    financial_model = self._create_model_from_comprehensive_data(comprehensive_facts, annual_data, quarterly_data, enhanced_fuzzy_matching, progress_callback)
                    
                    progress("    ✓ Comprehensive financial model created successfully!")
                    return financial_model
                    
                except Exception as e:
                    progress(f"    ✗ Error in comprehensive XBRL processing: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    return financial_model
        except Exception as e:
            progress(f"    ✗ Error creating financial model: {str(e)}")
            return financial_model

    def _extract_financial_data(self, facts: Dict, metrics: Dict, period: str) -> pd.DataFrame:
        """
        Extract financial data from SEC facts for a given period and set of metrics.
        Now includes validation to filter out non-financial data.
        
        Args:
            facts (Dict): SEC facts data
            metrics (Dict): Dictionary of metric names and their SEC tags
            period (str): 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        data = {}
        validation_stats = {
            'total_points': 0,
            'valid_points': 0,
            'rejected_points': 0
        }
        
        for metric_name, sec_tag in metrics.items():
            if sec_tag in facts:
                metric_data = facts[sec_tag]
                units = metric_data.get('units', {})
                
                # Find the appropriate unit (USD is most common)
                unit_key = None
                for key in units.keys():
                    if 'USD' in key or key == 'USD':
                        unit_key = key
                        break
                
                if unit_key:
                    periods = units[unit_key]
                    
                    # Filter by period
                    filtered_periods = []
                    for period_data in periods:
                        if period == 'annual' and period_data.get('form') in ['10-K', '10-K/A']:
                            filtered_periods.append(period_data)
                        elif period == 'quarterly' and period_data.get('form') in ['10-Q', '10-Q/A']:
                            filtered_periods.append(period_data)
                    
                    # Sort by end date (most recent first)
                    filtered_periods.sort(key=lambda x: x.get('end', ''), reverse=True)
                    
                    # Take the most recent periods based on the quarters parameter
                    # For annual data, limit to the number of years needed
                    # For quarterly data, limit to the actual quarters requested
                    if period == 'annual':
                        # For annual data, we want the most recent years
                        # If quarters=5, we want 2 years (5/4 rounded up)
                        years_needed = max(1, (self.current_quarters + 3) // 4)
                        recent_periods = filtered_periods[:years_needed]
                    else:
                        # For quarterly data, we want the most recent quarters
                        recent_periods = filtered_periods[:self.current_quarters]
                    
                    for period_data in recent_periods:
                        end_date = period_data.get('end', '')
                        value = period_data.get('val', 0)
                        
                        validation_stats['total_points'] += 1
                        
                        # Validate the data point
                        if self._validate_financial_data(metric_name, value, unit_key):
                            if end_date not in data:
                                data[end_date] = {}
                            data[end_date][metric_name] = value
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
        
        # Log validation statistics for SEC API path
        if hasattr(self, 'progress_callback') and self.progress_callback:
            self.progress_callback(f"    • SEC API validation: {validation_stats['valid_points']}/{validation_stats['total_points']} points passed validation")
            if validation_stats['rejected_points'] > 0:
                self.progress_callback(f"    • Rejected {validation_stats['rejected_points']} non-financial data points from SEC API")
        
        # Convert to DataFrame
        if data:
            df = pd.DataFrame.from_dict(data, orient='index')
            df.index = pd.to_datetime(df.index)
            df = df.sort_index()
            return df
        else:
            return pd.DataFrame()
    
    def create_sensitivity_model(self, financial_model: Dict[str, pd.DataFrame], 
                               ticker: str, quarters: int = 8, progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Create a sensitivity analysis model with operating leverage impacts.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            ticker (str): Stock ticker symbol
            quarters (int): Number of quarters used in the financial model (for reference)
            progress_callback (callable): Optional callback function for progress updates
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing sensitivity analysis dataframes
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            sensitivity_model = {
                'case_summary': pd.DataFrame(),
                'financial_model': pd.DataFrame(),
                'kpi_summary': pd.DataFrame()
            }
            
            # Get the most recent annual income statement for base calculations
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            
            if annual_income.empty:
                progress("    ✗ No annual income statement data available for sensitivity analysis")
                return sensitivity_model
            
            progress("    • Creating operating leverage scenarios...")
            # Create case summary sheet with operating leverage scenarios
            case_summary = self._create_case_summary(annual_income)
            sensitivity_model['case_summary'] = case_summary
            progress("    ✓ Operating leverage scenarios created")
            
            progress("    • Creating enhanced financial model with historical and forecasted data...")
            # Create enhanced financial model with historical and forecasted data
            enhanced_model = self._create_enhanced_financial_model(financial_model)
            sensitivity_model['financial_model'] = enhanced_model
            progress("    ✓ Enhanced financial model created")
            
            progress("    • Generating KPI summary sheet...")
            # Create KPI summary sheet
            kpi_summary = self._create_kpi_summary(financial_model)
            sensitivity_model['kpi_summary'] = kpi_summary
            progress("    ✓ KPI summary sheet created")
            
            return sensitivity_model
            
        except Exception as e:
            progress(f"    ✗ Error in create_sensitivity_model: {str(e)}")
            return {
                'case_summary': pd.DataFrame(),
                'financial_model': pd.DataFrame(),
                'kpi_summary': pd.DataFrame()
            }
    
    def _create_case_summary(self, annual_income: pd.DataFrame) -> pd.DataFrame:
        """
        Create case summary sheet showing operating leverage impacts.
        
        Args:
            annual_income (pd.DataFrame): Annual income statement data
            
        Returns:
            pd.DataFrame: Case summary with sensitivity scenarios
        """
        if annual_income.empty:
            return pd.DataFrame()
        
        # Get the most recent year's data
        latest_year = annual_income.index[-1]
        base_data = annual_income.loc[latest_year]
        
        # Define sensitivity scenarios
        scenarios = {
            'Base Case': 1.0,
            'Optimistic (+20%)': 1.2,
            'Pessimistic (-20%)': 0.8,
            'High Growth (+50%)': 1.5,
            'Recession (-30%)': 0.7
        }
        
        case_summary_data = []
        
        for scenario_name, revenue_multiplier in scenarios.items():
            # Calculate revenue impact
            base_revenue = base_data.get('Revenues', 0)
            new_revenue = base_revenue * revenue_multiplier
            
            # Calculate operating leverage impact
            # Assume fixed costs remain constant and variable costs scale with revenue
            base_cogs = base_data.get('CostOfRevenue', 0)
            base_opex = base_data.get('OperatingExpenses', 0)
            
            # Assume 70% of COGS is variable, 30% of OpEx is variable
            variable_cogs_ratio = 0.7
            variable_opex_ratio = 0.3
            
            new_cogs = (base_cogs * variable_cogs_ratio * revenue_multiplier + 
                       base_cogs * (1 - variable_cogs_ratio))
            new_opex = (base_opex * variable_opex_ratio * revenue_multiplier + 
                       base_opex * (1 - variable_opex_ratio))
            
            new_gross_profit = new_revenue - new_cogs
            new_operating_income = new_gross_profit - new_opex
            
            # Calculate operating leverage
            revenue_change = (new_revenue - base_revenue) / base_revenue if base_revenue != 0 else 0
            operating_income_change = (new_operating_income - base_data.get('OperatingIncomeLoss', 0)) / base_data.get('OperatingIncomeLoss', 0) if base_data.get('OperatingIncomeLoss', 0) != 0 else 0
            
            operating_leverage = operating_income_change / revenue_change if revenue_change != 0 else 0
            
            case_summary_data.append({
                'Scenario': scenario_name,
                'Revenue': new_revenue,
                'COGS': new_cogs,
                'Gross Profit': new_gross_profit,
                'Operating Expenses': new_opex,
                'Operating Income': new_operating_income,
                'Revenue Change %': revenue_change * 100,
                'Operating Income Change %': operating_income_change * 100,
                'Operating Leverage': operating_leverage
            })
        
        return pd.DataFrame(case_summary_data)
    
    def _create_enhanced_financial_model(self, financial_model: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Create enhanced financial model with historical and forecasted data.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            
        Returns:
            pd.DataFrame: Enhanced financial model
        """
        # Combine all financial data into one comprehensive model
        enhanced_data = {}
        
        # Add historical data
        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow']:
            for period in ['annual', 'quarterly']:
                key = f'{period}_{statement_type}'
                if key in financial_model and not financial_model[key].empty:
                    df = financial_model[key]
                    for date in df.index:
                        for column in df.columns:
                            enhanced_data[f"{period}_{statement_type}_{column}"] = {
                                'date': date,
                                'value': df.loc[date, column],
                                'type': 'historical',
                                'period': period,
                                'statement': statement_type
                            }
        
        # Add forecasted data (simple linear projection for demonstration)
        if enhanced_data:
            # Get the most recent dates for each period
            annual_dates = []
            quarterly_dates = []
            
            for key, data in enhanced_data.items():
                if data['period'] == 'annual':
                    annual_dates.append(data['date'])
                elif data['period'] == 'quarterly':
                    quarterly_dates.append(data['date'])
            
            if annual_dates:
                latest_annual = max(annual_dates)
                # Add 3 years of forecast
                for i in range(1, 4):
                    forecast_date = latest_annual + pd.DateOffset(years=i)
                    enhanced_data[f"forecast_annual_year_{i}"] = {
                        'date': forecast_date,
                        'value': 0,  # Placeholder for forecasted values
                        'type': 'forecast',
                        'period': 'annual',
                        'statement': 'projection'
                    }
            
            if quarterly_dates:
                latest_quarterly = max(quarterly_dates)
                # Add 4 quarters of forecast
                for i in range(1, 5):
                    forecast_date = latest_quarterly + pd.DateOffset(months=i*3)
                    enhanced_data[f"forecast_quarterly_q{i}"] = {
                        'date': forecast_date,
                        'value': 0,  # Placeholder for forecasted values
                        'type': 'forecast',
                        'period': 'quarterly',
                        'statement': 'projection'
                    }
        
        # Convert to DataFrame
        if enhanced_data:
            df = pd.DataFrame.from_dict(enhanced_data, orient='index')
            df = df.sort_values('date')
            return df
        else:
            return pd.DataFrame()
    
    def _create_kpi_summary(self, financial_model: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Create KPI summary sheet with key financial metrics.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            
        Returns:
            pd.DataFrame: KPI summary sheet
        """
        kpi_data = []
        
        # Calculate KPIs from available data
        annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
        annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
        annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
        
        if not annual_income.empty:
            latest_income = annual_income.iloc[-1]
            
            # Profitability KPIs
            revenue = latest_income.get('Revenues', 0)
            net_income = latest_income.get('NetIncomeLoss', 0)
            operating_income = latest_income.get('OperatingIncomeLoss', 0)
            gross_profit = latest_income.get('GrossProfit', 0)
            
            if revenue != 0:
                kpi_data.extend([
                    {'KPI': 'Net Profit Margin', 'Value': (net_income / revenue) * 100, 'Unit': '%'},
                    {'KPI': 'Operating Margin', 'Value': (operating_income / revenue) * 100, 'Unit': '%'},
                    {'KPI': 'Gross Margin', 'Value': (gross_profit / revenue) * 100, 'Unit': '%'}
                ])
        
        if not annual_balance.empty:
            latest_balance = annual_balance.iloc[-1]
            
            # Efficiency KPIs
            total_assets = latest_balance.get('TotalAssets', 0)
            total_equity = latest_balance.get('TotalStockholdersEquity', 0)
            
            if total_assets != 0:
                kpi_data.append({'KPI': 'Return on Assets (ROA)', 'Value': (net_income / total_assets) * 100, 'Unit': '%'})
            
            if total_equity != 0:
                kpi_data.append({'KPI': 'Return on Equity (ROE)', 'Value': (net_income / total_equity) * 100, 'Unit': '%'})
            
            # Liquidity KPIs
            current_assets = latest_balance.get('TotalCurrentAssets', 0)
            current_liabilities = latest_balance.get('TotalCurrentLiabilities', 0)
            
            if current_liabilities != 0:
                kpi_data.append({'KPI': 'Current Ratio', 'Value': current_assets / current_liabilities, 'Unit': 'x'})
        
        if not annual_cash_flow.empty:
            latest_cash_flow = annual_cash_flow.iloc[-1]
            
            # Cash Flow KPIs
            operating_cash_flow = latest_cash_flow.get('NetCashProvidedByUsedInOperatingActivities', 0)
            capital_expenditures = abs(latest_cash_flow.get('CapitalExpenditures', 0))
            
            if capital_expenditures != 0:
                kpi_data.append({'KPI': 'Operating Cash Flow to CapEx', 'Value': operating_cash_flow / capital_expenditures, 'Unit': 'x'})
        
        # Add growth metrics if multiple years available
        if len(annual_income) >= 2:
            current_revenue = annual_income.iloc[-1].get('Revenues', 0)
            previous_revenue = annual_income.iloc[-2].get('Revenues', 0)
            
            if previous_revenue != 0:
                revenue_growth = ((current_revenue - previous_revenue) / previous_revenue) * 100
                kpi_data.append({'KPI': 'Revenue Growth (YoY)', 'Value': revenue_growth, 'Unit': '%'})
        
        return pd.DataFrame(kpi_data)
    
    def get_filing_content(self, accession_number: str, primary_document: str) -> str:
        """
        Get the content of a specific SEC filing.
        
        Args:
            accession_number (str): SEC accession number
            primary_document (str): Primary document name
            
        Returns:
            str: Filing content
        """
        try:
            # Format accession number
            accession_number = accession_number.replace('-', '')
            
            # Construct URL for filing
            filing_url = f"https://www.sec.gov/Archives/edgar/data/{accession_number}/{primary_document}"
            
            response = self.session.get(filing_url)
            
            if response.status_code == 200:
                return response.text
            else:
                print(f"Error fetching filing content: {response.status_code}")
                return ""
                
        except Exception as e:
            print(f"Error in get_filing_content: {str(e)}")
            return ""
    
    def export_to_excel_fast_preview(self, financial_model: Dict[str, pd.DataFrame], 
                                   sensitivity_model: Dict[str, pd.DataFrame], 
                                   ticker: str, filename: str, progress_callback=None) -> str:
        """
        Export financial models to Excel file with minimal formatting for fast preview generation.
        This method skips the expensive professional formatting to provide quick results.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import pandas as pd
        import os
        from datetime import datetime
        
        progress("    • Creating fast preview Excel file...")
        
        # Ensure the Storage directory exists
        storage_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Storage')
        os.makedirs(storage_dir, exist_ok=True)
        
        filepath = os.path.join(storage_dir, filename)
        
        progress(f"    • Writing data to Excel: {filename}")

        # Write all DataFrames to Excel using pandas (fast, no formatting)
        progress("    • Writing financial data to Excel sheets...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Annual sheet
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
            annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
            if not annual_income.empty or not annual_balance.empty or not annual_cash_flow.empty:
                progress("    • Writing annual financial statements...")
                startrow = 0
                if not annual_income.empty:
                    annual_income.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_income.shape[1] + 3
                if not annual_balance.empty:
                    annual_balance.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_balance.shape[1] + 3
                if not annual_cash_flow.empty:
                    annual_cash_flow.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
            
            # Quarterly sheet
            quarterly_income = financial_model.get('quarterly_income_statement', pd.DataFrame())
            quarterly_balance = financial_model.get('quarterly_balance_sheet', pd.DataFrame())
            quarterly_cash_flow = financial_model.get('quarterly_cash_flow', pd.DataFrame())
            if not quarterly_income.empty or not quarterly_balance.empty or not quarterly_cash_flow.empty:
                progress("    • Writing quarterly financial statements...")
                startrow = 0
                if not quarterly_income.empty:
                    quarterly_income.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_income.shape[1] + 3
                if not quarterly_balance.empty:
                    quarterly_balance.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_balance.shape[1] + 3
                if not quarterly_cash_flow.empty:
                    quarterly_cash_flow.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
            
            # Sensitivity and summary sheets
            progress("    • Writing sensitivity analysis and summary sheets...")
            for sheet_name, df in sensitivity_model.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name.replace('_', ' ').title())
            
            # Summary sheet
            summary_data = {
                'Metric': ['Ticker', 'Model Created', 'Data Points', 'Scenarios Analyzed'],
                'Value': [
                    ticker,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sum(len(df) for df in financial_model.values() if not df.empty),
                    len(sensitivity_model.get('case_summary', pd.DataFrame()))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        progress(f"    ✓ Fast preview Excel file saved: {filepath}")
        return filepath

    def apply_excel_formatting(self, filepath: str, schmoove_mode: bool = False, progress_callback=None) -> str:
        """
        Apply professional formatting to an existing Excel file.
        This can be used to upgrade a fast preview file to full formatting.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        
        if not os.path.exists(filepath):
            progress(f"    ✗ File not found: {filepath}")
            return filepath
        
        progress("    • Applying professional formatting to existing Excel file...")
        
        if schmoove_mode:
            try:
                import joblib
                import psutil
                # Limit joblib to 75% of CPUs
                n_jobs = max(1, int(psutil.cpu_count(logical=True) * 0.75))
                progress(f"    • Schmoove mode enabled: Using {n_jobs} CPU cores")
            except ImportError:
                n_jobs = 1
        else:
            n_jobs = 1
        
        # Parent-child mapping for indentation following standard three-statement modeling
        parent_map = {
            # Income Statement - Standard format
            'Revenue': None,  # Top level
            'CostOfGoodsSold': 'Revenue',
            'GrossProfit': None,  # Calculated total
            
            # Operating expenses
            'ResearchAndDevelopmentExpense': 'OperatingExpenses',
            'SellingGeneralAndAdministrativeExpense': 'OperatingExpenses',
            'DepreciationAndAmortization': 'OperatingExpenses',
            'StockBasedCompensationExpense': 'OperatingExpenses',
            'RestructuringCharges': 'OperatingExpenses',
            'ImpairmentCharges': 'OperatingExpenses',
            'OtherOperatingExpenses': 'OperatingExpenses',
            'OperatingExpenses': None,  # Subtotal
            
            'OperatingIncome': None,  # Total
            
            # Non-operating items
            'InterestIncome': 'NonOperatingIncome',
            'InterestExpense': 'NonOperatingIncome',
            'GainLossOnSaleOfAssets': 'NonOperatingIncome',
            'ForeignCurrencyGainLoss': 'NonOperatingIncome',
            'OtherIncomeExpense': 'NonOperatingIncome',
            'NonOperatingIncome': None,  # Subtotal
            
            'IncomeBeforeTaxes': None,  # Total
            'IncomeTaxExpense': 'IncomeBeforeTaxes',
            'NetIncome': None,  # Final total
            
            # Earnings per share
            'EarningsPerShareBasic': 'NetIncome',
            'EarningsPerShareDiluted': 'NetIncome',
            'WeightedAverageSharesBasic': 'NetIncome',
            'WeightedAverageSharesDiluted': 'NetIncome',
            
            # Balance Sheet - Standard format: Assets = Liabilities + Equity
            # Current Assets
            'CashAndCashEquivalents': 'CurrentAssets',
            'ShortTermInvestments': 'CurrentAssets',
            'AccountsReceivable': 'CurrentAssets',
            'Inventory': 'CurrentAssets',
            'PrepaidExpenses': 'CurrentAssets',
            'OtherCurrentAssets': 'CurrentAssets',
            'CurrentAssets': None,  # Subtotal
            
            # Non-Current Assets
            'PropertyPlantAndEquipmentNet': 'NonCurrentAssets',
            'Goodwill': 'NonCurrentAssets',
            'IntangibleAssetsNet': 'NonCurrentAssets',
            'LongTermInvestments': 'NonCurrentAssets',
            'DeferredTaxAssets': 'NonCurrentAssets',
            'OtherLongTermAssets': 'NonCurrentAssets',
            'NonCurrentAssets': None,  # Subtotal
            
            'TotalAssets': None,  # Total Assets
            
            # Current Liabilities
            'AccountsPayable': 'CurrentLiabilities',
            'AccruedExpenses': 'CurrentLiabilities',
            'DeferredRevenue': 'CurrentLiabilities',
            'ShortTermDebt': 'CurrentLiabilities',
            'OtherCurrentLiabilities': 'CurrentLiabilities',
            'CurrentLiabilities': None,  # Subtotal
            
            # Non-Current Liabilities
            'LongTermDebt': 'NonCurrentLiabilities',
            'DeferredTaxLiabilities': 'NonCurrentLiabilities',
            'OtherLongTermLiabilities': 'NonCurrentLiabilities',
            'NonCurrentLiabilities': None,  # Subtotal
            
            'TotalLiabilities': None,  # Total Liabilities
            
            # Stockholders' Equity
            'CommonStock': 'StockholdersEquity',
            'AdditionalPaidInCapital': 'StockholdersEquity',
            'RetainedEarnings': 'StockholdersEquity',
            'AccumulatedOtherComprehensiveIncome': 'StockholdersEquity',
            'TreasuryStock': 'StockholdersEquity',
            'StockholdersEquity': None,  # Subtotal
            
            # Calculated metrics
            'WorkingCapital': None,
            'TotalDebt': None,
            
            # Cash Flow Statement - Standard format
            # Operating Activities
            'NetIncome': None,  # Starting point
            'DepreciationAndAmortization': 'OperatingAdjustments',
            'StockBasedCompensation': 'OperatingAdjustments',
            'DeferredIncomeTaxes': 'OperatingAdjustments',
            'OperatingAdjustments': None,  # Subtotal
            
            # Changes in Working Capital
            'ChangeInAccountsReceivable': 'WorkingCapitalChanges',
            'ChangeInInventory': 'WorkingCapitalChanges',
            'ChangeInAccountsPayable': 'WorkingCapitalChanges',
            'ChangeInDeferredRevenue': 'WorkingCapitalChanges',
            'ChangeInOtherWorkingCapital': 'WorkingCapitalChanges',
            'WorkingCapitalChanges': None,  # Subtotal
            
            'OtherOperatingActivities': 'OperatingActivities',
            'OperatingActivities': None,  # Subtotal
            'NetCashFromOperatingActivities': None,  # Total
            
            # Investing Activities
            'CapitalExpenditures': 'InvestingActivities',
            'Acquisitions': 'InvestingActivities',
            'Investments': 'InvestingActivities',
            'ProceedsFromInvestments': 'InvestingActivities',
            'OtherInvestingActivities': 'InvestingActivities',
            'InvestingActivities': None,  # Subtotal
            'NetCashFromInvestingActivities': None,  # Total
            
            # Financing Activities
            'ProceedsFromDebt': 'FinancingActivities',
            'RepaymentsOfDebt': 'FinancingActivities',
            'DividendsPaid': 'FinancingActivities',
            'StockRepurchases': 'FinancingActivities',
            'ProceedsFromStockIssuance': 'FinancingActivities',
            'OtherFinancingActivities': 'FinancingActivities',
            'FinancingActivities': None,  # Subtotal
            'NetCashFromFinancingActivities': None,  # Total
            
            # Net Change and Ending Balance
            'EffectOfExchangeRateChanges': None,
            'NetChangeInCash': None,  # Final total
            'CashAtBeginningOfPeriod': None,
            'CashAtEndOfPeriod': None
        }

        # Open the workbook
        wb = openpyxl.load_workbook(filepath)
        
        # Format annual and quarterly sheets
        for sheet in ['Annual Financial Statements', 'Quarterly Financial Statements']:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                # Bold headers
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True)
                # Indent and bold parent categories (if present)
                def format_row(row):
                    line_item_cell = row[0]
                    if line_item_cell.value in parent_map:
                        parent = parent_map[line_item_cell.value]
                        if parent is None:
                            line_item_cell.font = Font(bold=True)
                        else:
                            indent_level = 1
                            p = parent
                            while p:
                                indent_level += 1
                                p = parent_map.get(p)
                            line_item_cell.alignment = Alignment(indent=indent_level)
                
                rows = list(ws.iter_rows(min_row=2))
                if schmoove_mode and n_jobs > 1:
                    progress("    • Applying formatting with parallel processing...")
                    try:
                        from joblib import Parallel, delayed
                        Parallel(n_jobs=n_jobs)(delayed(format_row)(row) for row in rows)
                    except ImportError:
                        # Fallback to sequential processing
                        for row in rows:
                            format_row(row)
                else:
                    for row in rows:
                        format_row(row)
        
        # Save the formatted workbook
        progress("    • Saving formatted Excel file...")
        wb.save(filepath)
        progress(f"    ✓ Excel file formatted and saved: {filepath}")
        return filepath

    def export_to_excel(self, financial_model: Dict[str, pd.DataFrame], 
                       sensitivity_model: Dict[str, pd.DataFrame], 
                       ticker: str, filename: str = None, schmoove_mode: bool = False, progress_callback=None) -> str:
        """
        Export financial models to Excel file with professional formatting for annual and quarterly sheets.
        If schmoove_mode is True, use parallel processing for formatting and allow higher resource usage.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        from datetime import datetime
        
        progress("    • Initializing Excel export...")
        
        if schmoove_mode:
            try:
                import joblib
                import psutil
                # Set pandas and openpyxl to use more threads if possible
                pd.set_option('compute.use_numexpr', True)
                # Limit joblib to 75% of CPUs
                n_jobs = max(1, int(psutil.cpu_count(logical=True) * 0.75))
                # Limit memory usage to half of available RAM
                mem = psutil.virtual_memory()
                max_mem = int(mem.total * 0.5)
                progress(f"    • Schmoove mode enabled: Using {n_jobs} CPU cores")
            except ImportError:
                n_jobs = 1
                max_mem = None
        else:
            n_jobs = 1
            max_mem = None
        
        # Parent-child mapping for indentation following standard three-statement modeling
        parent_map = {
            # Income Statement - Standard format
            'Revenue': None,  # Top level
            'CostOfGoodsSold': 'Revenue',
            'GrossProfit': None,  # Calculated total
            
            # Operating expenses
            'ResearchAndDevelopmentExpense': 'OperatingExpenses',
            'SellingGeneralAndAdministrativeExpense': 'OperatingExpenses',
            'DepreciationAndAmortization': 'OperatingExpenses',
            'StockBasedCompensationExpense': 'OperatingExpenses',
            'RestructuringCharges': 'OperatingExpenses',
            'ImpairmentCharges': 'OperatingExpenses',
            'OtherOperatingExpenses': 'OperatingExpenses',
            'OperatingExpenses': None,  # Subtotal
            
            'OperatingIncome': None,  # Total
            
            # Non-operating items
            'InterestIncome': 'NonOperatingIncome',
            'InterestExpense': 'NonOperatingIncome',
            'GainLossOnSaleOfAssets': 'NonOperatingIncome',
            'ForeignCurrencyGainLoss': 'NonOperatingIncome',
            'OtherIncomeExpense': 'NonOperatingIncome',
            'NonOperatingIncome': None,  # Subtotal
            
            'IncomeBeforeTaxes': None,  # Total
            'IncomeTaxExpense': 'IncomeBeforeTaxes',
            'NetIncome': None,  # Final total
            
            # Earnings per share
            'EarningsPerShareBasic': 'NetIncome',
            'EarningsPerShareDiluted': 'NetIncome',
            'WeightedAverageSharesBasic': 'NetIncome',
            'WeightedAverageSharesDiluted': 'NetIncome',
            
            # Balance Sheet - Standard format: Assets = Liabilities + Equity
            # Current Assets
            'CashAndCashEquivalents': 'CurrentAssets',
            'ShortTermInvestments': 'CurrentAssets',
            'AccountsReceivable': 'CurrentAssets',
            'Inventory': 'CurrentAssets',
            'PrepaidExpenses': 'CurrentAssets',
            'OtherCurrentAssets': 'CurrentAssets',
            'CurrentAssets': None,  # Subtotal
            
            # Non-Current Assets
            'PropertyPlantAndEquipmentNet': 'NonCurrentAssets',
            'Goodwill': 'NonCurrentAssets',
            'IntangibleAssetsNet': 'NonCurrentAssets',
            'LongTermInvestments': 'NonCurrentAssets',
            'DeferredTaxAssets': 'NonCurrentAssets',
            'OtherLongTermAssets': 'NonCurrentAssets',
            'NonCurrentAssets': None,  # Subtotal
            
            'TotalAssets': None,  # Total Assets
            
            # Current Liabilities
            'AccountsPayable': 'CurrentLiabilities',
            'AccruedExpenses': 'CurrentLiabilities',
            'DeferredRevenue': 'CurrentLiabilities',
            'ShortTermDebt': 'CurrentLiabilities',
            'OtherCurrentLiabilities': 'CurrentLiabilities',
            'CurrentLiabilities': None,  # Subtotal
            
            # Non-Current Liabilities
            'LongTermDebt': 'NonCurrentLiabilities',
            'DeferredTaxLiabilities': 'NonCurrentLiabilities',
            'OtherLongTermLiabilities': 'NonCurrentLiabilities',
            'NonCurrentLiabilities': None,  # Subtotal
            
            'TotalLiabilities': None,  # Total Liabilities
            
            # Stockholders' Equity
            'CommonStock': 'StockholdersEquity',
            'AdditionalPaidInCapital': 'StockholdersEquity',
            'RetainedEarnings': 'StockholdersEquity',
            'AccumulatedOtherComprehensiveIncome': 'StockholdersEquity',
            'TreasuryStock': 'StockholdersEquity',
            'StockholdersEquity': None,  # Subtotal
            
            # Calculated metrics
            'WorkingCapital': None,
            'TotalDebt': None,
            
            # Cash Flow Statement - Standard format
            # Operating Activities
            'NetIncome': None,  # Starting point
            'DepreciationAndAmortization': 'OperatingAdjustments',
            'StockBasedCompensation': 'OperatingAdjustments',
            'DeferredIncomeTaxes': 'OperatingAdjustments',
            'OperatingAdjustments': None,  # Subtotal
            
            # Changes in Working Capital
            'ChangeInAccountsReceivable': 'WorkingCapitalChanges',
            'ChangeInInventory': 'WorkingCapitalChanges',
            'ChangeInAccountsPayable': 'WorkingCapitalChanges',
            'ChangeInDeferredRevenue': 'WorkingCapitalChanges',
            'ChangeInOtherWorkingCapital': 'WorkingCapitalChanges',
            'WorkingCapitalChanges': None,  # Subtotal
            
            'OtherOperatingActivities': 'OperatingActivities',
            'OperatingActivities': None,  # Subtotal
            'NetCashFromOperatingActivities': None,  # Total
            
            # Investing Activities
            'CapitalExpenditures': 'InvestingActivities',
            'Acquisitions': 'InvestingActivities',
            'Investments': 'InvestingActivities',
            'ProceedsFromInvestments': 'InvestingActivities',
            'OtherInvestingActivities': 'InvestingActivities',
            'InvestingActivities': None,  # Subtotal
            'NetCashFromInvestingActivities': None,  # Total
            
            # Financing Activities
            'ProceedsFromDebt': 'FinancingActivities',
            'RepaymentsOfDebt': 'FinancingActivities',
            'DividendsPaid': 'FinancingActivities',
            'StockRepurchases': 'FinancingActivities',
            'ProceedsFromStockIssuance': 'FinancingActivities',
            'OtherFinancingActivities': 'FinancingActivities',
            'FinancingActivities': None,  # Subtotal
            'NetCashFromFinancingActivities': None,  # Total
            
            # Net Change and Ending Balance
            'EffectOfExchangeRateChanges': None,
            'NetChangeInCash': None,  # Final total
            'CashAtBeginningOfPeriod': None,
            'CashAtEndOfPeriod': None
        }

        # Ensure the Storage directory exists
        storage_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Storage')
        os.makedirs(storage_dir, exist_ok=True)
        
        if filename is None:
            filename = f"{ticker}_financial_model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(storage_dir, filename)
        
        progress(f"    • Creating Excel file: {filename}")

        # Write all DataFrames to Excel using pandas (fast)
        progress("    • Writing financial data to Excel sheets...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Annual sheet
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
            annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
            if not annual_income.empty or not annual_balance.empty or not annual_cash_flow.empty:
                progress("    • Writing annual financial statements...")
                startrow = 0
                if not annual_income.empty:
                    annual_income.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_income.shape[1] + 3
                if not annual_balance.empty:
                    annual_balance.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_balance.shape[1] + 3
                if not annual_cash_flow.empty:
                    annual_cash_flow.transpose().to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
            # Quarterly sheet
            quarterly_income = financial_model.get('quarterly_income_statement', pd.DataFrame())
            quarterly_balance = financial_model.get('quarterly_balance_sheet', pd.DataFrame())
            quarterly_cash_flow = financial_model.get('quarterly_cash_flow', pd.DataFrame())
            if not quarterly_income.empty or not quarterly_balance.empty or not quarterly_cash_flow.empty:
                progress("    • Writing quarterly financial statements...")
                startrow = 0
                if not quarterly_income.empty:
                    quarterly_income.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_income.shape[1] + 3
                if not quarterly_balance.empty:
                    quarterly_balance.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_balance.shape[1] + 3
                if not quarterly_cash_flow.empty:
                    quarterly_cash_flow.transpose().to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
            # Sensitivity and summary sheets
            progress("    • Writing sensitivity analysis and summary sheets...")
            for sheet_name, df in sensitivity_model.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name.replace('_', ' ').title())
            # Summary sheet
            summary_data = {
                'Metric': ['Ticker', 'Model Created', 'Data Points', 'Scenarios Analyzed'],
                'Value': [
                    ticker,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sum(len(df) for df in financial_model.values() if not df.empty),
                    len(sensitivity_model.get('case_summary', pd.DataFrame()))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Now open with openpyxl and apply formatting
        progress("    • Applying professional formatting...")
        wb = openpyxl.load_workbook(filepath)
        # Format annual and quarterly sheets
        for sheet in ['Annual Financial Statements', 'Quarterly Financial Statements']:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                # Bold headers
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True)
                # Indent and bold parent categories (if present)
                def format_row(row):
                    line_item_cell = row[0]
                    if line_item_cell.value in parent_map:
                        parent = parent_map[line_item_cell.value]
                        if parent is None:
                            line_item_cell.font = Font(bold=True)
                        else:
                            indent_level = 1
                            p = parent
                            while p:
                                indent_level += 1
                                p = parent_map.get(p)
                            line_item_cell.alignment = Alignment(indent=indent_level)
                rows = list(ws.iter_rows(min_row=2))
                if schmoove_mode and n_jobs > 1:
                    progress("    • Applying formatting with parallel processing...")
                    Parallel(n_jobs=n_jobs)(delayed(format_row)(row) for row in rows)
                else:
                    for row in rows:
                        format_row(row)
        # Save the formatted workbook
        progress("    • Saving formatted Excel file...")
        wb.save(filepath)
        progress(f"    ✓ Excel file saved: {filepath}")
        return filepath

    def _create_vertically_stacked_statement(self, income_df, balance_df, cash_flow_df, period_type):
        """
        Create a vertically stacked financial statement with proper formatting following standard three-statement modeling principles.
        
        Args:
            income_df: Income statement DataFrame
            balance_df: Balance sheet DataFrame
            cash_flow_df: Cash flow statement DataFrame
            period_type: "Annual" or "Quarterly"
            
        Returns:
            pd.DataFrame: Combined vertically stacked statement
        """
        # Define the order of line items for each statement following standard accounting principles
        
        # INCOME STATEMENT - Standard format from top to bottom
        income_order = [
            # Revenue section
            'Revenue',
            'CostOfGoodsSold',
            'GrossProfit',
            
            # Operating expenses
            'ResearchAndDevelopmentExpense',
            'SellingGeneralAndAdministrativeExpense',
            'DepreciationAndAmortization',
            'StockBasedCompensationExpense',
            'RestructuringCharges',
            'ImpairmentCharges',
            'OtherOperatingExpenses',
            
            # Operating income
            'OperatingIncome',
            
            # Non-operating income/expense
            'InterestIncome',
            'InterestExpense',
            'GainLossOnSaleOfAssets',
            'ForeignCurrencyGainLoss',
            'OtherIncomeExpense',
            
            # Income before taxes
            'IncomeBeforeTaxes',
            
            # Income taxes
            'IncomeTaxExpense',
            
            # Net income
            'NetIncome',
            
            # Earnings per share
            'EarningsPerShareBasic',
            'EarningsPerShareDiluted',
            'WeightedAverageSharesBasic',
            'WeightedAverageSharesDiluted'
        ]
        
        # BALANCE SHEET - Standard format: Assets = Liabilities + Equity
        balance_order = [
            # Current Assets
            'CashAndCashEquivalents',
            'ShortTermInvestments',
            'AccountsReceivable',
            'Inventory',
            'PrepaidExpenses',
            'OtherCurrentAssets',
            'TotalCurrentAssets',
            
            # Non-Current Assets
            'PropertyPlantAndEquipmentNet',
            'Goodwill',
            'IntangibleAssetsNet',
            'LongTermInvestments',
            'DeferredTaxAssets',
            'OtherLongTermAssets',
            'TotalNonCurrentAssets',
            'TotalAssets',
            
            # Current Liabilities
            'AccountsPayable',
            'AccruedExpenses',
            'DeferredRevenue',
            'ShortTermDebt',
            'OtherCurrentLiabilities',
            'TotalCurrentLiabilities',
            
            # Non-Current Liabilities
            'LongTermDebt',
            'DeferredTaxLiabilities',
            'OtherLongTermLiabilities',
            'TotalNonCurrentLiabilities',
            'TotalLiabilities',
            
            # Stockholders' Equity
            'CommonStock',
            'AdditionalPaidInCapital',
            'RetainedEarnings',
            'AccumulatedOtherComprehensiveIncome',
            'TreasuryStock',
            'TotalStockholdersEquity',
            
            # Calculated metrics
            'WorkingCapital',
            'TotalDebt'
        ]
        
        # CASH FLOW STATEMENT - Standard format: Operating, Investing, Financing
        cash_flow_order = [
            # Operating Activities
            'NetIncome',
            'DepreciationAndAmortization',
            'StockBasedCompensation',
            'DeferredIncomeTaxes',
            
            # Changes in Working Capital
            'ChangeInAccountsReceivable',
            'ChangeInInventory',
            'ChangeInAccountsPayable',
            'ChangeInDeferredRevenue',
            'ChangeInOtherWorkingCapital',
            
            'OtherOperatingActivities',
            'NetCashFromOperatingActivities',
            
            # Investing Activities
            'CapitalExpenditures',
            'Acquisitions',
            'Investments',
            'ProceedsFromInvestments',
            'OtherInvestingActivities',
            'NetCashFromInvestingActivities',
            
            # Financing Activities
            'ProceedsFromDebt',
            'RepaymentsOfDebt',
            'DividendsPaid',
            'StockRepurchases',
            'ProceedsFromStockIssuance',
            'OtherFinancingActivities',
            'NetCashFromFinancingActivities',
            
            # Net Change and Ending Balance
            'EffectOfExchangeRateChanges',
            'NetChangeInCash',
            'CashAtBeginningOfPeriod',
            'CashAtEndOfPeriod'
        ]
        
        # Create combined DataFrame
        combined_data = {}
        
        # Add income statement items
        for item in income_order:
            if item in income_df.columns:
                for date in income_df.index:
                    key = f"INCOME STATEMENT - {item}"
                    if key not in combined_data:
                        combined_data[key] = {}
                    combined_data[key][date] = income_df.loc[date, item]
        
        # Add balance sheet items
        for item in balance_order:
            if item in balance_df.columns:
                for date in balance_df.index:
                    key = f"BALANCE SHEET - {item}"
                    if key not in combined_data:
                        combined_data[key] = {}
                    combined_data[key][date] = balance_df.loc[date, item]
        
        # Add cash flow statement items
        for item in cash_flow_order:
            if item in cash_flow_df.columns:
                for date in cash_flow_df.index:
                    key = f"CASH FLOW - {item}"
                    if key not in combined_data:
                        combined_data[key] = {}
                    combined_data[key][date] = cash_flow_df.loc[date, item]
        
        if combined_data:
            combined_df = pd.DataFrame.from_dict(combined_data, orient='index')
            combined_df = combined_df.sort_index()
            return combined_df
        else:
            return pd.DataFrame()

    def _extract_xbrl_from_filing(self, filing_row, cik, progress_callback=None):
        """
        Extract XBRL data from a specific filing.
        
        Args:
            filing_row: DataFrame row containing filing information
            cik: Company CIK number
            progress_callback: Optional callback function for progress updates
            
        Returns:
            Dict: Extracted XBRL facts
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            accession_number = filing_row['accessionNumber']
            cik_dir = str(int(cik))
            accession_clean = accession_number.replace('-', '')
            filing_dir_url = f"https://www.sec.gov/Archives/edgar/data/{cik_dir}/{accession_clean}/"
            
            # List files in the directory
            progress("        • Accessing filing directory...")
            dir_response = self.session.get(filing_dir_url)
            if dir_response.status_code != 200:
                progress(f"        ✗ Could not access filing directory: {filing_dir_url}")
                return None
            
            # Find .xml files (XBRL instance docs)
            import re
            xml_files = re.findall(r'href="([^"]+\.xml)"', dir_response.text)
            xbrl_instance_file = None
            
            progress(f"        • Found {len(xml_files)} XML files")
            
            # Prefer files ending with _htm.xml and not FilingSummary.xml
            for xml_file in xml_files:
                if xml_file.endswith('_htm.xml') and 'FilingSummary' not in xml_file:
                    xbrl_instance_file = xml_file
                    break
            
            if xbrl_instance_file is None:
                # Fallback: pick the first .xml file that's not FilingSummary
                for xml_file in xml_files:
                    if 'FilingSummary' not in xml_file:
                        xbrl_instance_file = xml_file
                        break
            
            if not xbrl_instance_file:
                progress(f"        ✗ No XBRL instance document found")
                return None
            
            # Construct URL
            if xbrl_instance_file.startswith('/'):
                xbrl_url = f"https://www.sec.gov{xbrl_instance_file}"
            else:
                xbrl_url = f"https://www.sec.gov/Archives/edgar/data/{cik_dir}/{accession_clean}/{xbrl_instance_file}"
            
            progress(f"        • Parsing XBRL: {xbrl_instance_file}")
            
            # Use Arelle to parse the XBRL instance document
            from arelle import Cntlr
            cntlr = Cntlr.Cntlr(logFileName=None)
            model_xbrl = cntlr.modelManager.load(xbrl_url)
            
            # Extract facts from the XBRL model
            facts_from_xml = {}
            fact_count = 0
            
            progress(f"        • Extracting facts from XBRL model...")
            for fact in model_xbrl.facts:
                try:
                    concept_name = str(fact.qname)
                    value = fact.value
                    context = fact.context
                    
                    if context is not None:
                        period = context.period
                        if period is not None:
                            end_date = getattr(period, 'endDate', None)
                            if end_date is None:
                                end_date = getattr(period, 'end', None)
                            if end_date is None:
                                end_date = getattr(context, 'endDate', None)
                            
                            if end_date:
                                if concept_name not in facts_from_xml:
                                    facts_from_xml[concept_name] = []
                                facts_from_xml[concept_name].append({
                                    'value': value,
                                    'end_date': end_date,
                                    'context': context,
                                    'filing_type': filing_row['form'],
                                    'filing_date': filing_row['filingDate']
                                })
                                fact_count += 1
                except Exception as e:
                    continue
            
            progress(f"        ✓ Extracted {len(facts_from_xml)} concepts ({fact_count} total facts)")
            return facts_from_xml
            
        except Exception as e:
            progress(f"        ✗ Error extracting XBRL from filing: {str(e)}")
            return None

    def _run_discrepancy_checks(self, annual_data, quarterly_data):
        """
        Run discrepancy checks between annual and quarterly data for key metrics.
        
        Args:
            annual_data: Annual data from 10-K filings
            quarterly_data: Quarterly data from 10-Q filings
        """
        print("\n" + "="*60)
        print("DISCREPANCY CHECKS: Annual vs Quarterly Data")
        print("="*60)
        
        # Key metrics to check
        key_metrics = [
            'us-gaap:NetIncomeLoss',
            'us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax',
            'us-gaap:GrossProfit',
            'us-gaap:OperatingIncomeLoss'
        ]
        
        for metric in key_metrics:
            if metric in annual_data and metric in quarterly_data:
                print(f"\nChecking: {metric}")
                
                # Get annual values
                annual_values = {}
                for item in annual_data[metric]:
                    date = item['end_date']
                    value = float(item['value'])
                    annual_values[date] = value
                
                # Get quarterly values and sum them by year
                quarterly_sums = {}
                for item in quarterly_data[metric]:
                    date = item['end_date']
                    year = date.year
                    value = float(item['value'])
                    
                    if year not in quarterly_sums:
                        quarterly_sums[year] = 0
                    quarterly_sums[year] += value
                
                # Compare annual vs quarterly sums
                for year, quarterly_sum in quarterly_sums.items():
                    # Find annual value for this year
                    annual_value = None
                    for date, value in annual_values.items():
                        if date.year == year:
                            annual_value = value
                            break
                    
                    if annual_value is not None:
                        difference = abs(annual_value - quarterly_sum)
                        difference_pct = (difference / annual_value * 100) if annual_value != 0 else 0
                        
                        print(f"  {year}: Annual={annual_value:,.0f}, Quarterly Sum={quarterly_sum:,.0f}")
                        print(f"    Difference: {difference:,.0f} ({difference_pct:.2f}%)")
                        
                        if difference_pct > 5:  # Flag differences > 5%
                            print(f"    SIGNIFICANT DISCREPANCY DETECTED!")
                        elif difference_pct > 1:  # Flag differences > 1%
                            print(f"    Minor discrepancy detected")
                        else:
                            print(f"    Data consistent")
            else:
                print(f"\nSkipping {metric}: Not available in both annual and quarterly data")

    def _create_model_from_comprehensive_data(self, comprehensive_facts, annual_data, quarterly_data, enhanced_fuzzy_matching=True, progress_callback=None):
        """
        Create financial model from comprehensive 10-K and 10-Q data.
        
        Args:
            comprehensive_facts: Combined facts from all filings (used for fuzzy matching)
            annual_data: Annual data from 10-K filings
            quarterly_data: Quarterly data from 10-Q filings
            enhanced_fuzzy_matching: Whether to include non-GAAP to GAAP mapping
            progress_callback: Optional callback function for progress updates
            
        Returns:
            Dict: Financial model with annual and quarterly data
        """
        # Define progress function for this method
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        financial_model = {
            'annual_income_statement': pd.DataFrame(),
            'annual_balance_sheet': pd.DataFrame(),
            'annual_cash_flow': pd.DataFrame(),
            'quarterly_income_statement': pd.DataFrame(),
            'quarterly_balance_sheet': pd.DataFrame(),
            'quarterly_cash_flow': pd.DataFrame()
        }
        
        # Get all available concepts from comprehensive_facts for fuzzy matching
        available_concepts = list(comprehensive_facts.keys())
        print(f"\nUsing fuzzy matching to find {len(available_concepts)} available concepts...")
        if enhanced_fuzzy_matching:
            print("Enhanced fuzzy matching enabled: Will include non-GAAP to GAAP mapping")
            progress("    • Enhanced fuzzy matching enabled (includes non-GAAP to GAAP mapping)")
        else:
            print("Standard fuzzy matching: GAAP concepts only")
            progress("    • Standard fuzzy matching enabled (GAAP concepts only)")
        
        progress("    • Applying standard three-statement financial modeling principles...")
        
        # Define desired metrics with multiple possible tags for each concept
        # Following standard three-statement financial modeling principles
        
        # CORE GAAP CONCEPTS - Essential line items that should always be included
        # These are the fundamental building blocks of financial statements
        
        # INCOME STATEMENT - Core GAAP concepts (top priority)
        core_income_statement_metrics = {
            # Revenue (top line)
            'Revenue': ['us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax', 'us-gaap:Revenues', 'us-gaap:SalesRevenueNet'],
            
            # Cost of goods sold
            'CostOfGoodsSold': ['us-gaap:CostOfRevenue', 'us-gaap:CostOfGoodsAndServicesSold'],
            
            # Gross profit
            'GrossProfit': ['us-gaap:GrossProfit', 'us-gaap:GrossProfitLoss'],
            
            # Core operating expenses
            'ResearchAndDevelopmentExpense': ['us-gaap:ResearchAndDevelopmentExpense'],
            'SellingGeneralAndAdministrativeExpense': ['us-gaap:SellingGeneralAndAdministrativeExpense'],
            
            # Operating income
            'OperatingIncome': ['us-gaap:OperatingIncomeLoss'],
            
            # Interest and taxes
            'InterestExpense': ['us-gaap:InterestExpense'],
            'IncomeTaxExpense': ['us-gaap:IncomeTaxExpenseBenefit'],
            
            # Net income (bottom line)
            'NetIncome': ['us-gaap:NetIncomeLoss'],
            
            # Earnings per share
            'EarningsPerShareBasic': ['us-gaap:EarningsPerShareBasic'],
            'EarningsPerShareDiluted': ['us-gaap:EarningsPerShareDiluted']
        }
        
        # BALANCE SHEET - Core GAAP concepts
        core_balance_sheet_metrics = {
            # Core current assets
            'CashAndCashEquivalents': ['us-gaap:CashAndCashEquivalentsAtCarryingValue'],
            'AccountsReceivable': ['us-gaap:AccountsReceivableNetCurrent'],
            'Inventory': ['us-gaap:InventoryNet'],
            'TotalCurrentAssets': ['us-gaap:AssetsCurrent'],
            
            # Core non-current assets
            'PropertyPlantAndEquipmentNet': ['us-gaap:PropertyPlantAndEquipmentNet'],
            'Goodwill': ['us-gaap:Goodwill'],
            'TotalAssets': ['us-gaap:Assets'],
            
            # Core current liabilities
            'AccountsPayable': ['us-gaap:AccountsPayableCurrent'],
            'TotalCurrentLiabilities': ['us-gaap:LiabilitiesCurrent'],
            
            # Core non-current liabilities
            'LongTermDebt': ['us-gaap:LongTermDebtNoncurrent'],
            'TotalLiabilities': ['us-gaap:Liabilities'],
            
            # Core equity
            'RetainedEarnings': ['us-gaap:RetainedEarningsAccumulatedDeficit'],
            'TotalStockholdersEquity': ['us-gaap:StockholdersEquity']
        }
        
        # CASH FLOW STATEMENT - Core GAAP concepts
        core_cash_flow_metrics = {
            # Operating activities
            'NetIncome': ['us-gaap:NetIncomeLoss'],
            'DepreciationAndAmortization': ['us-gaap:DepreciationAndAmortization'],
            'NetCashFromOperatingActivities': ['us-gaap:NetCashProvidedByUsedInOperatingActivities'],
            
            # Investing activities
            'CapitalExpenditures': ['us-gaap:PaymentsToAcquirePropertyPlantAndEquipment'],
            'NetCashFromInvestingActivities': ['us-gaap:NetCashProvidedByUsedInInvestingActivities'],
            
            # Financing activities
            'DividendsPaid': ['us-gaap:PaymentsOfDividends'],
            'NetCashFromFinancingActivities': ['us-gaap:NetCashProvidedByUsedInFinancingActivities'],
            
            # Net change
            'NetChangeInCash': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect']
        }
        
        # ADDITIONAL GAAP CONCEPTS - Up to 20 additional items that are GAAP-compliant and properly categorized
        additional_income_statement_metrics = {
            # Additional operating items
            'DepreciationAndAmortization': ['us-gaap:DepreciationAndAmortization'],
            'StockBasedCompensationExpense': ['us-gaap:StockBasedCompensationExpense'],
            'RestructuringCharges': ['us-gaap:RestructuringCharges'],
            'ImpairmentCharges': ['us-gaap:ImpairmentCharges'],
            
            # Additional non-operating items
            'InterestIncome': ['us-gaap:InterestIncome'],
            'GainLossOnSaleOfAssets': ['us-gaap:GainLossOnSaleOfAssets'],
            'ForeignCurrencyGainLoss': ['us-gaap:ForeignCurrencyGainLoss'],
            'OtherIncomeExpense': ['us-gaap:OtherIncomeExpenseNet'],
            
            # Additional metrics
            'IncomeBeforeTaxes': ['us-gaap:IncomeLossFromContinuingOperationsBeforeIncomeTaxes'],
            'WeightedAverageSharesBasic': ['us-gaap:WeightedAverageNumberOfSharesOutstandingBasic'],
            'WeightedAverageSharesDiluted': ['us-gaap:WeightedAverageNumberOfSharesOutstandingDiluted']
        }
        
        additional_balance_sheet_metrics = {
            # Additional current assets
            'ShortTermInvestments': ['us-gaap:ShortTermInvestments'],
            'PrepaidExpenses': ['us-gaap:PrepaidExpenseAndOtherAssetsCurrent'],
            'OtherCurrentAssets': ['us-gaap:OtherAssetsCurrent'],
            
            # Additional non-current assets
            'IntangibleAssetsNet': ['us-gaap:IntangibleAssetsNetExcludingGoodwill'],
            'LongTermInvestments': ['us-gaap:InvestmentsNoncurrent'],
            'DeferredTaxAssets': ['us-gaap:DeferredTaxAssetsNet'],
            'OtherLongTermAssets': ['us-gaap:OtherAssetsNoncurrent'],
            'TotalNonCurrentAssets': ['us-gaap:AssetsNoncurrent'],
            
            # Additional current liabilities
            'AccruedExpenses': ['us-gaap:AccruedLiabilitiesCurrent'],
            'DeferredRevenue': ['us-gaap:ContractWithCustomerLiabilityCurrent'],
            'ShortTermDebt': ['us-gaap:ShortTermBorrowings'],
            'OtherCurrentLiabilities': ['us-gaap:OtherLiabilitiesCurrent'],
            
            # Additional non-current liabilities
            'DeferredTaxLiabilities': ['us-gaap:DeferredTaxLiabilitiesNet'],
            'OtherLongTermLiabilities': ['us-gaap:OtherLiabilitiesNoncurrent'],
            'TotalNonCurrentLiabilities': ['us-gaap:LiabilitiesNoncurrent'],
            
            # Additional equity
            'CommonStock': ['us-gaap:CommonStockValue'],
            'AdditionalPaidInCapital': ['us-gaap:AdditionalPaidInCapital'],
            'AccumulatedOtherComprehensiveIncome': ['us-gaap:AccumulatedOtherComprehensiveIncomeLossNetOfTax'],
            'TreasuryStock': ['us-gaap:TreasuryStockValue'],
            
            # Additional calculated metrics
            'WorkingCapital': ['us-gaap:WorkingCapital']
        }
        
        additional_cash_flow_metrics = {
            # Additional operating activities
            'StockBasedCompensation': ['us-gaap:StockBasedCompensationExpense'],
            'DeferredIncomeTaxes': ['us-gaap:DeferredIncomeTaxExpenseBenefit'],
            'ChangeInAccountsReceivable': ['us-gaap:IncreaseDecreaseInAccountsReceivable'],
            'ChangeInInventory': ['us-gaap:IncreaseDecreaseInInventories'],
            'ChangeInAccountsPayable': ['us-gaap:IncreaseDecreaseInAccountsPayable'],
            'ChangeInDeferredRevenue': ['us-gaap:IncreaseDecreaseInContractWithCustomerLiability'],
            'OtherOperatingActivities': ['us-gaap:OtherOperatingActivitiesCashFlowStatement'],
            
            # Additional investing activities
            'Acquisitions': ['us-gaap:PaymentsToAcquireBusinessesNetOfCashAcquired'],
            'Investments': ['us-gaap:PaymentsToAcquireInvestments'],
            'ProceedsFromInvestments': ['us-gaap:ProceedsFromSaleMaturityAndCollectionsOfInvestments'],
            'OtherInvestingActivities': ['us-gaap:OtherInvestingActivitiesCashFlowStatement'],
            
            # Additional financing activities
            'ProceedsFromDebt': ['us-gaap:ProceedsFromIssuanceOfLongTermDebt'],
            'RepaymentsOfDebt': ['us-gaap:RepaymentsOfLongTermDebt'],
            'StockRepurchases': ['us-gaap:PaymentsForRepurchaseOfCommonStock'],
            'ProceedsFromStockIssuance': ['us-gaap:ProceedsFromIssuanceOfCommonStock'],
            'OtherFinancingActivities': ['us-gaap:OtherFinancingActivitiesCashFlowStatement'],
            
            # Additional metrics
            'EffectOfExchangeRateChanges': ['us-gaap:EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents'],
            'CashAtBeginningOfPeriod': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodBeginningBalance'],
            'CashAtEndOfPeriod': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodEndingBalance']
        }
        
        # Function to selectively add additional metrics (up to 20 total additional items)
        def add_selective_additional_metrics(core_mapping, additional_metrics, available_concepts, enhanced_fuzzy_matching):
            """Add up to 20 additional GAAP-compliant metrics that are available in the data."""
            additional_mapping = {}
            total_additional = 0
            max_additional = 20
            
            progress(f"    • Adding up to {max_additional} additional GAAP-compliant line items...")
            
            for concept_name, possible_tags in additional_metrics.items():
                if total_additional >= max_additional:
                    break
                    
                for tag in possible_tags:
                    # Check if this concept is available and not already in core mapping
                    if tag in available_concepts and concept_name not in core_mapping:
                        mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                        if mapping:
                            additional_mapping.update(mapping)
                            total_additional += 1
                            progress(f"      ✓ Added: {concept_name}")
                            break
            
            progress(f"    • Added {total_additional} additional line items")
            return additional_mapping
        
        # Start with core metrics
        progress("    • Matching core GAAP concepts...")
        
        # Match core income statement concepts
        income_mapping = {}
        for concept_name, possible_tags in core_income_statement_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    income_mapping.update(mapping)
                    break
        
        # Match core balance sheet concepts
        balance_mapping = {}
        for concept_name, possible_tags in core_balance_sheet_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    balance_mapping.update(mapping)
                    break
        
        # Match core cash flow concepts
        cash_flow_mapping = {}
        for concept_name, possible_tags in core_cash_flow_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    cash_flow_mapping.update(mapping)
                    break
        
        # Add selective additional metrics
        income_mapping.update(add_selective_additional_metrics(income_mapping, additional_income_statement_metrics, available_concepts, enhanced_fuzzy_matching))
        balance_mapping.update(add_selective_additional_metrics(balance_mapping, additional_balance_sheet_metrics, available_concepts, enhanced_fuzzy_matching))
        cash_flow_mapping.update(add_selective_additional_metrics(cash_flow_mapping, additional_cash_flow_metrics, available_concepts, enhanced_fuzzy_matching))
        
        # Create annual dataframes (from annual_data - 10-K filings)
        print(f"\nCreating annual financial statements...")
        for statement, mapping in [('income_statement', income_mapping),
                                 ('balance_sheet', balance_mapping),
                                 ('cash_flow', cash_flow_mapping)]:
            
            df = self._create_dataframe_from_data(annual_data, mapping, 'annual')
            financial_model[f'annual_{statement}'] = df
            print(f"Annual {statement}: {len(df)} rows extracted")
        
        # Create quarterly dataframes (from quarterly_data - 10-Q filings)
        print(f"\nCreating quarterly financial statements...")
        for statement, mapping in [('income_statement', income_mapping),
                                 ('balance_sheet', balance_mapping),
                                 ('cash_flow', cash_flow_mapping)]:
            
            df = self._create_dataframe_from_data(quarterly_data, mapping, 'quarterly')
            financial_model[f'quarterly_{statement}'] = df
            print(f"Quarterly {statement}: {len(df)} rows extracted")
        
        return financial_model

    def _create_dataframe_from_data(self, data, mapping, period_type):
        """
        Create DataFrame from annual or quarterly data using concept mapping.
        Now includes comprehensive data validation to filter out non-financial data.
        
        Args:
            data: Annual or quarterly data dictionary
            mapping: Dictionary mapping concept names to actual XBRL tags
            period_type: 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        return self._extract_and_validate_financial_data(data, mapping, period_type)

    def _extract_and_validate_financial_data(self, data, mapping, period_type):
        """
        Extract financial data with comprehensive validation to filter out non-financial data.
        
        Args:
            data: Annual or quarterly data dictionary
            mapping: Dictionary mapping concept names to actual XBRL tags
            period_type: 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        df_data = {}
        mapped_tags = set(mapping.values())
        
        # Collect all data points first, then filter by quarters
        all_data_points = []
        validation_stats = {
            'total_points': 0,
            'valid_points': 0,
            'rejected_points': 0,
            'rejected_reasons': {}
        }
        
        # 1. Add mapped concepts with validation
        for concept_name, actual_tag in mapping.items():
            if actual_tag in data:
                for item in data[actual_tag]:
                    validation_stats['total_points'] += 1
                    date = item['end_date']
                    
                    try:
                        value = float(item['value'])
                        unit = item.get('unit', '')
                        
                        # Validate the data point
                        if self._validate_financial_data(concept_name, value, unit):
                            all_data_points.append({
                                'date': date,
                                'concept': concept_name,
                                'value': value
                            })
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
                            reason = f"Non-financial data: {concept_name} = {value}"
                            validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                            
                    except (ValueError, TypeError):
                        validation_stats['rejected_points'] += 1
                        reason = f"Invalid numeric value: {concept_name} = {item.get('value', 'N/A')}"
                        validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                        continue
        
        # 2. Add unmapped tags as their own columns (with validation)
        for tag in data:
            if tag not in mapped_tags:
                for item in data[tag]:
                    validation_stats['total_points'] += 1
                    date = item['end_date']
                    
                    try:
                        value = float(item['value'])
                        unit = item.get('unit', '')
                        
                        # Validate the data point
                        if self._validate_financial_data(tag, value, unit):
                            all_data_points.append({
                                'date': date,
                                'concept': tag,
                                'value': value
                            })
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
                            reason = f"Non-financial data: {tag} = {value}"
                            validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                            
                    except (ValueError, TypeError):
                        validation_stats['rejected_points'] += 1
                        reason = f"Invalid numeric value: {tag} = {item.get('value', 'N/A')}"
                        validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                        continue
        
        # Log validation statistics
        if hasattr(self, 'progress_callback') and self.progress_callback:
            self.progress_callback(f"    • Data validation: {validation_stats['valid_points']}/{validation_stats['total_points']} points passed validation")
            if validation_stats['rejected_points'] > 0:
                self.progress_callback(f"    • Rejected {validation_stats['rejected_points']} non-financial data points")
                # Log top rejection reasons
                top_reasons = sorted(validation_stats['rejected_reasons'].items(), key=lambda x: x[1], reverse=True)[:3]
                for reason, count in top_reasons:
                    self.progress_callback(f"      - {reason}: {count} occurrences")
        
        if all_data_points:
            # Sort by date (most recent first)
            all_data_points.sort(key=lambda x: x['date'], reverse=True)
            
            # Filter to requested number of periods
            if period_type == 'annual':
                # For annual data, limit to the number of years needed
                years_needed = max(1, (self.current_quarters + 3) // 4)
                # Get unique years and take the most recent ones
                unique_years = []
                for point in all_data_points:
                    year = point['date'].year
                    if year not in unique_years:
                        unique_years.append(year)
                        if len(unique_years) >= years_needed:
                            break
                
                # Filter data points to only include the selected years
                filtered_points = [point for point in all_data_points if point['date'].year in unique_years]
            else:
                # For quarterly data, limit to the requested number of quarters
                # Get unique quarters and take the most recent ones
                unique_quarters = []
                for point in all_data_points:
                    quarter_key = (point['date'].year, point['date'].month)
                    if quarter_key not in unique_quarters:
                        unique_quarters.append(quarter_key)
                        if len(unique_quarters) >= self.current_quarters:
                            break
                
                # Filter data points to only include the selected quarters
                filtered_points = [point for point in all_data_points if (point['date'].year, point['date'].month) in unique_quarters]
            
            # Convert filtered data to DataFrame format
            for point in filtered_points:
                date_str = point['date'].strftime('%Y-%m-%d')
                if date_str not in df_data:
                    df_data[date_str] = {}
                df_data[date_str][point['concept']] = point['value']
        
        if df_data:
            df = pd.DataFrame.from_dict(df_data, orient='index')
            df.index = pd.to_datetime(df.index)
            df = df.sort_index()
            return df
        else:
            return pd.DataFrame()

    def _fuzzy_match_concepts(self, desired_concepts: Dict[str, str], available_concepts: List[str], 
                             similarity_threshold: float = 0.85, enhanced_fuzzy_matching: bool = True) -> Dict[str, str]:
        """
        Use fuzzy matching to find the best matches between desired XBRL concepts and available concepts.
        Also, for any non-US GAAP available concept, map it to the closest US GAAP tag if similarity > 0.7.
        Limited to 20 highest scoring mappings per financial statement section.
        Optimized: uses rapidfuzz if available, caches results, and parallelizes matching.
        
        Args:
            desired_concepts: Dictionary of concept names to desired XBRL tags
            available_concepts: List of available XBRL concepts
            similarity_threshold: Minimum similarity score for matching (default: 0.85)
            enhanced_fuzzy_matching: Whether to include non-GAAP to GAAP mapping (default: True)
            
        Returns:
            Dict[str, str]: Mapping of concept names to matched XBRL tags
        """
        concept_mapping = {}
        us_gaap_tags = [c for c in available_concepts if str(c).startswith('us-gaap:')]
        non_us_gaap_tags = [c for c in available_concepts if not str(c).startswith('us-gaap:')]
        cache_key = (tuple(sorted(desired_concepts.items())), tuple(sorted(available_concepts)), enhanced_fuzzy_matching)
        if cache_key in self._fuzzy_match_cache:
            return self._fuzzy_match_cache[cache_key].copy()
        def match_one(concept_name, desired_tag):
            best_match = None
            best_score = 0
            desired_main = desired_tag.split(':')[-1] if ':' in desired_tag else desired_tag
            if RAPIDFUZZ_AVAILABLE:
                # Use rapidfuzz for fast similarity
                matches = rf_process.extract(desired_main, [str(ac).split(':')[-1] for ac in available_concepts], scorer=fuzz.ratio, limit=3)
                for match_main, score, idx in matches:
                    score = score / 100.0
                    available_concept_str = str(available_concepts[idx])
                    if available_concept_str == desired_tag:
                        return (concept_name, available_concept_str, 1.0)
                    if score > best_score:
                        best_score = score
                        best_match = available_concept_str
            else:
                for available_concept in available_concepts:
                    available_concept_str = str(available_concept)
                    available_main = available_concept_str.split(':')[-1] if ':' in available_concept_str else available_concept_str
                    if available_concept_str == desired_tag:
                        return (concept_name, available_concept_str, 1.0)
                    if available_main == desired_main:
                        return (concept_name, available_concept_str, 0.95)
                    similarity = SequenceMatcher(None, available_main.lower(), desired_main.lower()).ratio()
                    if desired_main.lower() in available_main.lower() or available_main.lower() in desired_main.lower():
                        similarity = max(similarity, 0.8)
                    if similarity > best_score:
                        best_score = similarity
                        best_match = available_concept_str
            return (concept_name, best_match, best_score)
        # Parallelize matching
        results = Parallel(n_jobs=-1)(delayed(match_one)(concept_name, desired_tag) for concept_name, desired_tag in desired_concepts.items())
        for concept_name, best_match, best_score in results:
            if best_score >= similarity_threshold and best_match:
                concept_mapping[concept_name] = best_match
        
        # Only include non-GAAP to GAAP mapping if enhanced_fuzzy_matching is enabled
        if enhanced_fuzzy_matching:
            # --- New: For each non-US GAAP tag, map to closest US GAAP tag if similarity > 0.7 ---
            def match_non_gaap(non_gaap):
                non_gaap_str = str(non_gaap)
                non_gaap_main = non_gaap_str.split(':')[-1] if ':' in non_gaap_str else non_gaap_str
                best_usgaap = None
                best_score = 0
                if RAPIDFUZZ_AVAILABLE:
                    matches = rf_process.extract(non_gaap_main, [str(ug).split(':')[-1] for ug in us_gaap_tags], scorer=fuzz.ratio, limit=3)
                    for match_main, score, idx in matches:
                        score = score / 100.0
                        usgaap_str = str(us_gaap_tags[idx])
                        if score > best_score:
                            best_score = score
                            best_usgaap = usgaap_str
                else:
                    for usgaap in us_gaap_tags:
                        usgaap_str = str(usgaap)
                        usgaap_main = usgaap_str.split(':')[-1] if ':' in usgaap_str else usgaap_str
                        similarity = SequenceMatcher(None, non_gaap_main.lower(), usgaap_main.lower()).ratio()
                        if non_gaap_main.lower() in usgaap_main.lower() or usgaap_main.lower() in non_gaap_main.lower():
                            similarity = max(similarity, 0.8)
                        if similarity > best_score:
                            best_score = similarity
                            best_usgaap = usgaap_str
                return (non_gaap_str, best_usgaap, best_score)
            potential_mappings = Parallel(n_jobs=-1)(delayed(match_non_gaap)(non_gaap) for non_gaap in non_us_gaap_tags)
            potential_mappings = [x for x in potential_mappings if x[1] and x[2] > 0.7]
            potential_mappings.sort(key=lambda x: x[2], reverse=True)
            top_mappings = potential_mappings[:20]
            for non_gaap_str, best_usgaap, score in top_mappings:
                concept_mapping[non_gaap_str] = best_usgaap
        
        self._fuzzy_match_cache[cache_key] = concept_mapping.copy()
        return concept_mapping

    def get_quarter_configurations(self) -> Dict[str, Dict[str, int]]:
        """
        Get predefined quarter configurations for different analysis periods.
        
        Returns:
            Dict[str, Dict[str, int]]: Dictionary of configuration names to their details
        """
        return {
            "short_term": {
                "quarters": 4,
                "years": 1,
                "description": "1 year of data - good for recent performance analysis"
            },
            "medium_term": {
                "quarters": 8,
                "years": 2,
                "description": "2 years of data - balanced view for most analyses"
            },
            "long_term": {
                "quarters": 12,
                "years": 3,
                "description": "3 years of data - good for trend analysis"
            },
            "extended": {
                "quarters": 16,
                "years": 4,
                "description": "4 years of data - comprehensive historical view"
            },
            "maximum": {
                "quarters": 20,
                "years": 5,
                "description": "5 years of data - maximum recommended for performance"
            }
        }
    
    def print_quarter_configurations(self):
        """
        Print available quarter configurations with descriptions.
        """
        configs = self.get_quarter_configurations()
        print("\nAvailable Quarter Configurations:")
        print("=" * 60)
        for name, details in configs.items():
            print(f"{name:12} : {details['quarters']:2d} quarters ({details['years']} years) - {details['description']}")
        print("=" * 60)
        print("Usage: create_financial_model(ticker, quarters=configs['medium_term']['quarters'])")

    def _validate_financial_data(self, concept_name: str, value: float, unit: str = None) -> bool:
        """
        Validate that a data point represents actual financial amounts and not other types of data.
        
        Args:
            concept_name: Name of the financial concept
            value: The numeric value to validate
            unit: The unit of measurement (if available)
            
        Returns:
            bool: True if the data appears to be valid financial data, False otherwise
        """
        # Convert to string for pattern matching
        value_str = str(value).lower()
        concept_lower = concept_name.lower()
        
        # 1. Check for year/date patterns
        year_patterns = [
            r'^20\d{2}$',  # Years like 2023, 2024
            r'^19\d{2}$',  # Years like 1999, 2000
            r'^\d{4}$'     # Any 4-digit number that could be a year
        ]
        
        for pattern in year_patterns:
            if re.match(pattern, value_str):
                return False
        
        # 2. Check for employee counts and headcount data
        employee_indicators = [
            'employee', 'headcount', 'personnel', 'staff', 'workforce',
            'full-time', 'part-time', 'fte', 'head count'
        ]
        
        for indicator in employee_indicators:
            if indicator in concept_lower:
                return False
        
        # 3. Check for percentage values (should be excluded from financial statements)
        percentage_indicators = [
            'percentage', 'percent', 'rate', 'ratio', 'margin', 'pct',
            'growth rate', 'return on', 'roi', 'roe', 'roa'
        ]
        
        for indicator in percentage_indicators:
            if indicator in concept_lower:
                return False
        
        # 4. Check for unit-based validation
        if unit:
            unit_lower = unit.lower()
            # Reject non-monetary units
            non_monetary_units = [
                'shares', 'units', 'employees', 'people', 'customers',
                'locations', 'stores', 'facilities', 'countries',
                'percent', 'percentage', 'ratio', 'times', 'days'
            ]
            
            for non_unit in non_monetary_units:
                if non_unit in unit_lower:
                    return False
        
        # 5. Check for suspicious value ranges
        # Reject very small numbers that are likely percentages (0.01 to 0.99)
        if 0.01 <= abs(value) <= 0.99:
            # But allow if it's clearly a financial amount (millions, billions)
            if abs(value) >= 1000000:  # Allow millions+
                pass
            else:
                # Check if concept suggests it should be a percentage
                percentage_concepts = ['margin', 'ratio', 'rate', 'return', 'growth']
                if any(pc in concept_lower for pc in percentage_concepts):
                    return False
        
        # 6. Check for count-based concepts
        count_indicators = [
            'number of', 'count of', 'total number', 'quantity',
            'shares outstanding', 'common stock', 'preferred stock'
        ]
        
        for indicator in count_indicators:
            if indicator in concept_lower:
                # Allow share counts but reject other counts
                if 'share' in indicator and ('outstanding' in indicator or 'stock' in indicator):
                    pass  # Allow share counts
                else:
                    return False
        
        # 7. Check for text-like values
        if isinstance(value, str) and not value.replace('.', '').replace('-', '').isdigit():
            return False
        
        # 8. Check for extreme values that might be errors
        if abs(value) > 1e15:  # Values over 1 quadrillion are suspicious
            return False
        
        # 9. Check for common non-financial concepts
        non_financial_indicators = [
            'age', 'duration', 'length', 'width', 'height', 'weight',
            'temperature', 'speed', 'distance', 'area', 'volume',
            'efficiency', 'productivity', 'satisfaction', 'score'
        ]
        
        for indicator in non_financial_indicators:
            if indicator in concept_lower:
                return False
        
        return True

# Example usage and testing
if __name__ == "__main__":
    sourcer = SECFileSourcer()
    
    # Show available configurations
    sourcer.print_quarter_configurations()
    
    # Use command-line arguments if provided, otherwise default to AAPL
    if len(sys.argv) > 1:
        tickers = [arg.upper() for arg in sys.argv[1:]]
    else:
        tickers = ["AAPL"]  # Default to Apple for testing

    # Get quarter configurations
    configs = sourcer.get_quarter_configurations()
    
    # Use medium_term (8 quarters) as default, but users can modify this
    quarters_to_use = configs["medium_term"]["quarters"]
    
    print(f"\nUsing {quarters_to_use} quarters of data ({quarters_to_use//4} years)")

    for ticker in tickers:
        print(f"\n{'='*40}\nTesting for ticker: {ticker}\n{'='*40}")
        print(f"Finding SEC filings for {ticker}...")
        filings = sourcer.find_sec_filings(ticker)
        print(f"Found {len(filings)} filings")
        if not filings.empty:
            print("Recent filings:")
            print(filings.head())
        print(f"\nCreating financial model for {ticker} with {quarters_to_use} quarters of data...")
        financial_model = sourcer.create_financial_model(ticker, quarters=quarters_to_use)
        if any(not df.empty for df in financial_model.values()):
            print(f"\nCreating sensitivity analysis for {ticker}...")
            sensitivity_model = sourcer.create_sensitivity_model(financial_model, ticker, quarters=quarters_to_use)
            excel_file = sourcer.export_to_excel(financial_model, sensitivity_model, ticker)
            print(f"\nModel creation complete! Check the Excel file: {excel_file}")
        else:
            print("\nNo financial data was successfully pulled. Excel file will not be created.") 